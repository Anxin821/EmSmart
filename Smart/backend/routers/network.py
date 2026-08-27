"""
智能工厂工作任务管理平台 - 车间网络路由
提供服务器/老化架/WiFi AP CRUD、状态检测、批量导入导出
"""
import subprocess
import platform
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook
from datetime import datetime

from core.database import get_db
from core.auth import get_current_user, require_role, write_operation_log
from schemas import (
    ServerCreate, ServerUpdate, AgingRackCreate, AgingRackUpdate,
    WifiApCreate, WifiApUpdate, ApiResponse, PaginatedData,
)
from core.crud import (
    get_servers_paginated, create_server, update_server, delete_server,
    get_aging_racks_paginated, create_aging_rack, update_aging_rack, delete_aging_rack,
    get_wifi_aps_paginated, create_wifi_ap, update_wifi_ap, delete_wifi_ap,
)

router = APIRouter(prefix="/network", tags=["车间网络"])


def _ping_device(ip: str) -> bool:
    """
    使用系统 ping 检测设备是否在线，禁止 shell 注入
    """
    if not ip or not isinstance(ip, str):
        return False
    ip = ip.strip()
    try:
        import socket
        socket.inet_aton(ip)  # 先校验确实是合法 IPv4，无效会抛异常
    except (OSError, ValueError):
        return False
    try:
        if platform.system().lower() == "windows":
            result = subprocess.run(
                ["ping", "-n", "1", "-w", "2000", ip],
                capture_output=True, timeout=5, shell=False
            )
        else:
            result = subprocess.run(
                ["ping", "-c", "1", "-W", "2", ip],
                capture_output=True, timeout=5, shell=False
            )
        return result.returncode == 0
    except Exception:
        return False


# ============================================================
# 服务器
# ============================================================
@router.get("/servers")
def list_servers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, total = get_servers_paginated(db, page, page_size, keyword, production_line, status)
    return ApiResponse(data=PaginatedData(
        total=total, page=page, page_size=page_size,
        items=[_server_to_dict(d) for d in items],
    ))


@router.post("/servers")
def add_server(
    data: ServerCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    server = create_server(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "server", data.server_id, f"新增服务器: {data.name}", request)
    return ApiResponse(data=_server_to_dict(server))


@router.put("/servers/{server_id}")
def edit_server(
    server_id: str,
    data: ServerUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    server = update_server(db, server_id, data.model_dump(exclude_unset=True))
    if not server:
        raise HTTPException(status_code=404, detail="服务器不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "server", server_id, "更新服务器", request)
    return ApiResponse(data=_server_to_dict(server))


@router.delete("/servers/{server_id}")
def remove_server(
    server_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    ok = delete_server(db, server_id)
    if not ok:
        raise HTTPException(status_code=404, detail="服务器不存在")
    write_operation_log(db, current_user["username"], "DELETE", "server", server_id, "删除服务器", request)
    return ApiResponse(message="删除成功")


@router.post("/servers/check-all")
def check_all_servers(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """手动触发全部服务器心跳检测"""
    from models import Server
    servers = db.query(Server).all()
    online = offline = 0
    for s in servers:
        if s.ip_address:
            is_alive = _ping_device(s.ip_address)
            s.status = "在线" if is_alive else "离线"
            s.last_check_time = datetime.utcnow()
            if is_alive:
                online += 1
            else:
                offline += 1
    db.commit()
    write_operation_log(db, current_user["username"], "UPDATE", "server", None, f"心跳检测: {online}在线 {offline}离线", request)
    return ApiResponse(data={"online": online, "offline": offline})


@router.post("/servers/import")
def import_servers(file: UploadFile = File(...), request: Request = None,
                   db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin", "engineer"))):
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls")
    wb = load_workbook(filename=BytesIO(file.file.read()), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    field_map = {"服务器ID": "server_id", "名称": "name", "产线": "production_line",
                 "机柜位置": "rack_location", "IP": "ip_address", "型号": "model",
                 "OS": "os", "状态": "status", "CPU使用率": "cpu_usage",
                 "内存使用率": "memory_usage", "硬盘使用率": "disk_usage", "负责人": "responsible_person"}
    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        r = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                f = field_map.get(str(h), str(h))
                r[f] = row[i]
        if r.get("server_id"):
            count += 1
            create_server(db, r)
    wb.close()
    write_operation_log(db, current_user["username"], "CREATE", "server", None, f"批量导入 {count} 台服务器", request)
    return ApiResponse(data={"imported": count})


@router.get("/servers/export")
def export_servers(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_servers_paginated(db, 1, 10000, keyword, production_line, status)
    wb = Workbook(); ws = wb.active; ws.title = "服务器"
    ws.append(["服务器ID","名称","产线","机柜位置","IP","型号","OS","状态","CPU%","内存%","硬盘%","负责人","最后检测"])
    for d in items:
        ws.append([d.server_id, d.name, d.production_line, d.rack_location, d.ip_address,
                    d.model, d.os, d.status, d.cpu_usage, d.memory_usage, d.disk_usage,
                    d.responsible_person, str(d.last_check_time)])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=servers.xlsx"})


# ============================================================
# 老化架
# ============================================================
@router.get("/aging-racks")
def list_aging_racks(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = get_aging_racks_paginated(db, page, page_size, keyword, production_line, status)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=[_rack_to_dict(d) for d in items]))


@router.post("/aging-racks")
def add_aging_rack(data: AgingRackCreate, request: Request, db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin", "engineer"))):
    rack = create_aging_rack(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "aging_rack", data.rack_id, f"新增老化架: {data.name}", request)
    return ApiResponse(data=_rack_to_dict(rack))


@router.put("/aging-racks/{rack_id}")
def edit_aging_rack(rack_id: str, data: AgingRackUpdate, request: Request,
                    db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    rack = update_aging_rack(db, rack_id, data.model_dump(exclude_unset=True))
    if not rack: raise HTTPException(404, "老化架不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "aging_rack", rack_id, "更新老化架", request)
    return ApiResponse(data=_rack_to_dict(rack))


@router.delete("/aging-racks/{rack_id}")
def remove_aging_rack(rack_id: str, request: Request, db: Session = Depends(get_db),
                      current_user: dict = Depends(require_role("admin"))):
    if not delete_aging_rack(db, rack_id): raise HTTPException(404, "老化架不存在")
    write_operation_log(db, current_user["username"], "DELETE", "aging_rack", rack_id, "删除老化架", request)
    return ApiResponse(message="删除成功")


@router.get("/aging-racks/export")
def export_aging_racks(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_aging_racks_paginated(db, 1, 10000, keyword, production_line, status)
    wb = Workbook(); ws = wb.active; ws.title = "老化架"
    ws.append(["老化架ID","名称","产线","位置","IP","总槽位","在用槽位","状态","负责人"])
    for d in items:
        ws.append([d.rack_id, d.name, d.production_line, d.location, d.ip_address, d.total_slots, d.used_slots, d.status, d.responsible_person])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=aging_racks.xlsx"})


# ============================================================
# WiFi AP
# ============================================================
@router.get("/wifi-aps")
def list_wifi_aps(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = get_wifi_aps_paginated(db, page, page_size, keyword, production_line, status)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=[_ap_to_dict(d) for d in items]))


@router.post("/wifi-aps")
def add_wifi_ap(data: WifiApCreate, request: Request, db: Session = Depends(get_db),
                current_user: dict = Depends(require_role("admin", "engineer"))):
    ap = create_wifi_ap(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "wifi_ap", data.ap_id, f"新增AP: {data.ssid}", request)
    return ApiResponse(data=_ap_to_dict(ap))


@router.put("/wifi-aps/{ap_id}")
def edit_wifi_ap(ap_id: str, data: WifiApUpdate, request: Request,
                 db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    ap = update_wifi_ap(db, ap_id, data.model_dump(exclude_unset=True))
    if not ap: raise HTTPException(404, "AP不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "wifi_ap", ap_id, "更新AP", request)
    return ApiResponse(data=_ap_to_dict(ap))


@router.delete("/wifi-aps/{ap_id}")
def remove_wifi_ap(ap_id: str, request: Request, db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin"))):
    if not delete_wifi_ap(db, ap_id): raise HTTPException(404, "AP不存在")
    write_operation_log(db, current_user["username"], "DELETE", "wifi_ap", ap_id, "删除AP", request)
    return ApiResponse(message="删除成功")


@router.get("/wifi-aps/export")
def export_wifi_aps(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_wifi_aps_paginated(db, 1, 10000, keyword, production_line, status)
    wb = Workbook(); ws = wb.active; ws.title = "WiFi AP"
    ws.append(["AP_ID","SSID","产线","IP","位置","信道","连接设备","状态","负责人"])
    for d in items:
        ws.append([d.ap_id, d.ssid, d.production_line, d.ip_address, d.location, d.channel, d.connected_devices, d.status, d.responsible_person])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=wifi_aps.xlsx"})


# ============================================================
# 辅助函数
# ============================================================
def _server_to_dict(d) -> dict:
    return {
        "id": d.id, "server_id": d.server_id, "name": d.name,
        "production_line": d.production_line, "rack_location": d.rack_location,
        "ip_address": d.ip_address, "model": d.model, "os": d.os,
        "status": d.status, "cpu_usage": d.cpu_usage, "memory_usage": d.memory_usage,
        "disk_usage": d.disk_usage, "responsible_person": d.responsible_person,
        "last_check_time": d.last_check_time.isoformat() if d.last_check_time else None,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _rack_to_dict(d) -> dict:
    return {
        "id": d.id, "rack_id": d.rack_id, "name": d.name,
        "production_line": d.production_line, "location": d.location,
        "ip_address": d.ip_address, "total_slots": d.total_slots,
        "used_slots": d.used_slots, "status": d.status,
        "responsible_person": d.responsible_person,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _ap_to_dict(d) -> dict:
    return {
        "id": d.id, "ap_id": d.ap_id, "ssid": d.ssid,
        "production_line": d.production_line, "ip_address": d.ip_address,
        "location": d.location, "channel": d.channel,
        "connected_devices": d.connected_devices, "status": d.status,
        "responsible_person": d.responsible_person,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }
