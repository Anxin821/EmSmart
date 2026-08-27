"""
智能工厂工作任务管理平台 - AOI&AI 设备路由
提供设备 CRUD、批量导入导出、选项查询
"""
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from core.database import get_db
from core.auth import get_current_user, require_role, check_module_access, write_operation_log
from schemas import AoiAiDeviceCreate, AoiAiDeviceUpdate, ApiResponse, PaginatedData
from core.crud import (
    get_aoi_devices_paginated, create_aoi_device, update_aoi_device,
    delete_aoi_device, get_aoi_device_by_id, get_aoi_device_by_pk, batch_import_devices,
)

router = APIRouter(prefix="/devices", tags=["AOI&AI设备"])


@router.get("")
def list_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取设备分页列表，支持按产线/状态/类型筛选和关键字搜索"""
    items, total = get_aoi_devices_paginated(db, page, page_size, keyword, production_line, status, device_type)
    return ApiResponse(data=PaginatedData(
        total=total, page=page, page_size=page_size,
        items=[_device_to_dict(d) for d in items],
    ))


@router.post("")
def create_device(
    data: AoiAiDeviceCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """新增设备（需 admin 或 engineer 权限）"""
    existing = get_aoi_device_by_id(db, data.device_id)
    if existing:
        raise HTTPException(status_code=400, detail="设备ID已存在")
    device = create_aoi_device(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "aoi_device", data.device_id, f"新增设备: {data.name}", request)
    return ApiResponse(data=_device_to_dict(device))


@router.put("/{item_id}")
def edit_device(
    item_id: int,
    data: AoiAiDeviceUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """编辑设备"""
    device = update_aoi_device(db, item_id, data.model_dump(exclude_unset=True))
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "aoi_device", str(item_id), f"更新设备", request)
    return ApiResponse(data=_device_to_dict(device))


@router.delete("/{item_id}")
def remove_device(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    """删除设备（仅管理员）"""
    ok = delete_aoi_device(db, item_id)
    if not ok:
        raise HTTPException(status_code=404, detail="设备不存在")
    write_operation_log(db, current_user["username"], "DELETE", "aoi_device", str(item_id), "删除设备", request)
    return ApiResponse(message="删除成功")


@router.get("/detail/{item_id}")
def get_device(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取单个设备详情"""
    device = get_aoi_device_by_pk(db, item_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return ApiResponse(data=_device_to_dict(device))


@router.post("/import")
def import_devices(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """批量导入设备（Excel）"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents), read_only=True)
    ws = wb.active

    # 解析表头行
    headers = [cell.value for cell in ws[1]]
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                # 映射中文表头到英文字段
                field_map = {
                    "设备ID": "device_id", "名称": "name", "类型": "device_type",
                    "产线": "production_line", "位置": "location", "IP": "ip_address",
                    "状态": "status", "负责人": "responsible_person", "安装日期": "install_date",
                    "备注": "remark",
                }
                field = field_map.get(str(h), str(h))
                row_dict[field] = row[i]
        if row_dict.get("device_id"):
            rows_data.append(row_dict)
    wb.close()

    count = batch_import_devices(db, rows_data)
    write_operation_log(db, current_user["username"], "CREATE", "aoi_device", None, f"批量导入 {count} 台设备", request)
    return ApiResponse(data={"imported": count}, message=f"成功导入 {count} 条记录")


@router.get("/export/excel")
def export_devices(
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导出设备列表为 Excel"""
    items, _ = get_aoi_devices_paginated(db, 1, 10000, keyword, production_line, status, device_type)

    wb = Workbook()
    ws = wb.active
    ws.title = "AOI&AI设备"
    headers = ["设备ID", "名称", "类型", "产线", "位置", "IP地址", "状态", "负责人", "安装日期", "备注"]
    ws.append(headers)

    for d in items:
        ws.append([
            d.device_id, d.name, d.device_type, d.production_line, d.location or "",
            d.ip_address or "", d.status or "", d.responsible_person or "",
            str(d.install_date) if d.install_date else "", d.remark or "",
        ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=aoi_ai_devices.xlsx"},
    )


def _clean_device_status(status) -> str:
    """统一清洗设备状态：去除两端的 | 与空白字符，空值返回 '正常'。"""
    if status is None:
        return "正常"
    s = str(status).strip().lstrip("|").rstrip("|").strip()
    return s or "正常"


def _device_to_dict(d) -> dict:
    """将 ORM 对象转为字典"""
    return {
        "id": d.id, "device_id": d.device_id, "name": d.name,
        "device_type": d.device_type, "production_line": d.production_line,
        "location": d.location, "ip_address": d.ip_address,
        "status": _clean_device_status(d.status),
        "responsible_person": d.responsible_person,
        "install_date": d.install_date.isoformat() if d.install_date else None,
        "remark": d.remark,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }
