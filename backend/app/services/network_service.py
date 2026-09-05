import platform
import subprocess
from io import BytesIO
from typing import Optional, Dict, Any, List
from datetime import datetime
from openpyxl import Workbook, load_workbook
from fastapi import HTTPException
from sqlalchemy.orm import Session

from app.repositories import network_repository as repo
from app.core.crud import write_operation_log
from app.core.timeutil import beijing_now


def _ping_device(ip: str) -> bool:
    if not ip or not isinstance(ip, str):
        return False
    ip = ip.strip()
    try:
        import socket
        socket.inet_aton(ip)
    except (OSError, ValueError):
        return False
    try:
        if platform.system().lower() == "windows":
            result = subprocess.run(["ping", "-n", "1", "-w", "2000", ip], capture_output=True, timeout=5, shell=False)
        else:
            result = subprocess.run(["ping", "-c", "1", "-W", "2", ip], capture_output=True, timeout=5, shell=False)
        return result.returncode == 0
    except Exception:
        return False


def _server_to_dict(d) -> Dict[str, Any]:
    return {
        "id": d.id, "server_id": d.server_id, "name": d.name,
        "production_line": d.production_line, "rack_location": d.rack_location,
        "ip_address": d.ip_address, "model": d.model, "os": d.os,
        "status": d.status, "cpu_usage": d.cpu_usage, "memory_usage": d.memory_usage,
        "disk_usage": d.disk_usage, "responsible_person": d.responsible_person,
        "last_check_time": d.last_check_time.isoformat() if d.last_check_time else None,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _rack_to_dict(d) -> Dict[str, Any]:
    return {
        "id": d.id, "rack_id": d.rack_id, "name": d.name,
        "production_line": d.production_line, "location": d.location,
        "ip_address": d.ip_address, "total_slots": d.total_slots,
        "used_slots": d.used_slots, "status": d.status,
        "responsible_person": d.responsible_person,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _ap_to_dict(d) -> Dict[str, Any]:
    return {
        "id": d.id, "ap_id": d.ap_id, "ssid": d.ssid,
        "production_line": d.production_line, "ip_address": d.ip_address,
        "location": d.location, "channel": d.channel,
        "connected_devices": d.connected_devices, "status": d.status,
        "responsible_person": d.responsible_person,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


# Servers

def list_servers(db: Session, page: int =1, page_size: int =20, keyword: Optional[str]=None, production_line: Optional[str]=None, status: Optional[str]=None):
    items, total = repo.list_servers(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ([_server_to_dict(d) for d in items], total)


def add_server(db: Session, data: dict, request, username: str):
    server = repo.create_server_repo(db, data)
    write_operation_log(db, username, "CREATE", "server", data.get("server_id"), f"新增服务器: {data.get('name')}", request)
    return _server_to_dict(server)


def edit_server(db: Session, server_id: str, data: dict, request, username: str):
    server = repo.update_server_repo(db, server_id, data)
    if not server:
        return None
    write_operation_log(db, username, "UPDATE", "server", server_id, "更新服务器", request)
    return _server_to_dict(server)


def remove_server(db: Session, server_id: str, request, username: str):
    ok = repo.delete_server_repo(db, server_id)
    if not ok:
        return False
    write_operation_log(db, username, "DELETE", "server", server_id, "删除服务器", request)
    return True


def check_all_servers(db: Session, request, username: str):
    from app.models import Server
    servers = db.query(Server).all()
    online = offline = 0
    for s in servers:
        if s.ip_address:
            is_alive = _ping_device(s.ip_address)
            s.status = "在线" if is_alive else "离线"
            s.last_check_time = beijing_now()
            if is_alive:
                online += 1
            else:
                offline += 1
    db.commit()
    write_operation_log(db, username, "UPDATE", "server", None, f"心跳检测: {online}在线 {offline}离线", request)
    return {"online": online, "offline": offline}


def import_servers(db: Session, rows: List[dict], request, username: str):
    count = 0
    for r in rows:
        if r.get("server_id"):
            repo.create_server_repo(db, r)
            count += 1
    write_operation_log(db, username, "CREATE", "server", None, f"批量导入 {count} 台服务器", request)
    return count


def export_servers_rows(db: Session, items):
    wb = Workbook(); ws = wb.active; ws.title = "服务器"
    ws.append(["服务器ID","名称","产线","机柜位置","IP","型号","OS","状态","CPU%","内存%","硬盘%","负责人","最后检测"])
    for d in items:
        ws.append([d.get("server_id"), d.get("name"), d.get("production_line"), d.get("rack_location"), d.get("ip_address"),
                    d.get("model"), d.get("os"), d.get("status"), d.get("cpu_usage"), d.get("memory_usage"), d.get("disk_usage"),
                    d.get("responsible_person"), str(d.get("last_check_time"))])
    output = BytesIO(); wb.save(output); output.seek(0)
    return output

# Aging racks

def list_aging_racks(db: Session, page: int =1, page_size: int =20, keyword: Optional[str]=None, production_line: Optional[str]=None, status: Optional[str]=None):
    items, total = repo.list_aging_racks(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ([_rack_to_dict(d) for d in items], total)


def add_aging_rack(db: Session, data: dict, request, username: str):
    rack = repo.create_aging_rack_repo(db, data)
    write_operation_log(db, username, "CREATE", "aging_rack", data.get("rack_id"), f"新增老化架: {data.get('name')}", request)
    return _rack_to_dict(rack)


def edit_aging_rack(db: Session, rack_id: str, data: dict, request, username: str):
    rack = repo.update_aging_rack_repo(db, rack_id, data)
    if not rack:
        return None
    write_operation_log(db, username, "UPDATE", "aging_rack", rack_id, "更新老化架", request)
    return _rack_to_dict(rack)


def remove_aging_rack(db: Session, rack_id: str, request, username: str):
    ok = repo.delete_aging_rack_repo(db, rack_id)
    if not ok:
        return False
    write_operation_log(db, username, "DELETE", "aging_rack", rack_id, "删除老化架", request)
    return True


def export_aging_racks_rows(db: Session, items):
    wb = Workbook(); ws = wb.active; ws.title = "老化架"
    ws.append(["老化架ID","名称","产线","位置","IP","总槽位","在用槽位","状态","负责人"])
    for d in items:
        ws.append([d.get("rack_id"), d.get("name"), d.get("production_line"), d.get("location"), d.get("ip_address"), d.get("total_slots"), d.get("used_slots"), d.get("status"), d.get("responsible_person")])
    output = BytesIO(); wb.save(output); output.seek(0)
    return output

# WiFi APs

def list_wifi_aps(db: Session, page: int =1, page_size: int =20, keyword: Optional[str]=None, production_line: Optional[str]=None, status: Optional[str]=None):
    items, total = repo.list_wifi_aps(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ([_ap_to_dict(d) for d in items], total)


def add_wifi_ap(db: Session, data: dict, request, username: str):
    ap = repo.create_wifi_ap_repo(db, data)
    write_operation_log(db, username, "CREATE", "wifi_ap", data.get("ap_id"), f"新增AP: {data.get('ssid')}", request)
    return _ap_to_dict(ap)


def edit_wifi_ap(db: Session, ap_id: str, data: dict, request, username: str):
    ap = repo.update_wifi_ap_repo(db, ap_id, data)
    if not ap:
        return None
    write_operation_log(db, username, "UPDATE", "wifi_ap", ap_id, "更新AP", request)
    return _ap_to_dict(ap)


def remove_wifi_ap(db: Session, ap_id: str, request, username: str):
    ok = repo.delete_wifi_ap_repo(db, ap_id)
    if not ok:
        return False
    write_operation_log(db, username, "DELETE", "wifi_ap", ap_id, "删除AP", request)
    return True


def export_wifi_aps_rows(db: Session, items):
    wb = Workbook(); ws = wb.active; ws.title = "WiFi AP"
    ws.append(["AP_ID","SSID","产线","IP","位置","信道","连接设备","状态","负责人"])
    for d in items:
        ws.append([d.get("ap_id"), d.get("ssid"), d.get("production_line"), d.get("ip_address"), d.get("location"), d.get("channel"), d.get("connected_devices"), d.get("status"), d.get("responsible_person")])
    output = BytesIO(); wb.save(output); output.seek(0)
    return output