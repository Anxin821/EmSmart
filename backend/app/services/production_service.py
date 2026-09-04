from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import HTTPException, UploadFile
from sqlalchemy.orm import Session

from app.repositories.production_repository import (
    batch_import_weekly,
    create_monthly_production,
    create_weekly_production,
    delete_monthly_production,
    delete_weekly_production,
    generate_monthly_from_weekly,
    get_monthly_production_paginated,
    get_monthly_summary_stats,
    get_monthly_trend,
    get_weekly_production_by_id,
    get_weekly_production_paginated,
    update_monthly_production,
    update_weekly_production,
)
from app.models import MonthlyProduction, WeeklyProduction
from app.schemas import MonthlyGenerateRequest, MonthlyProductionCreate, MonthlyProductionUpdate, WeeklyProductionCreate, WeeklyProductionUpdate


def weekly_to_dict(d) -> dict:
    return {
        "id": d.id, "year": d.year, "week_number": d.week_number,
        "production_line": d.production_line, "project": d.project,
        "total_output": d.total_output, "qualified_count": d.qualified_count,
        "defect_count": d.defect_count, "yield_rate": d.yield_rate,
        "recorder": d.recorder, "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def monthly_to_dict(d) -> dict:
    return {
        "id": d.id, "year": d.year, "month": d.month,
        "production_line": d.production_line, "project": d.project,
        "monthly_total_output": d.monthly_total_output,
        "monthly_qualified_count": d.monthly_qualified_count,
        "monthly_defect_count": d.monthly_defect_count,
        "monthly_yield_rate": d.monthly_yield_rate,
        "recorder": d.recorder, "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def list_weekly(db: Session, page: int, page_size: int, year: Optional[int] = None, week: Optional[int] = None, production_line: Optional[str] = None, project: Optional[str] = None):
    items, total = get_weekly_production_paginated(db, page, page_size, year, week, production_line, project)
    return [weekly_to_dict(item) for item in items], total


def get_weekly_by_id(db: Session, record_id: int):
    record = get_weekly_production_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    return weekly_to_dict(record)


def create_weekly(db: Session, payload: WeeklyProductionCreate):
    record = create_weekly_production(db, payload.model_dump())
    return weekly_to_dict(record)


def update_weekly(db: Session, record_id: int, payload: WeeklyProductionUpdate):
    record = update_weekly_production(db, record_id, payload.model_dump(exclude_unset=True))
    if not record:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    return weekly_to_dict(record)


def delete_weekly(db: Session, record_id: int):
    ok = delete_weekly_production(db, record_id)
    if not ok:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    return True


def import_weekly_excel(db: Session, file: UploadFile):
    from openpyxl import load_workbook

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")
    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    field_map = {
        "年": "year", "周数": "week_number", "产线": "production_line", "项目": "project",
        "总产量": "total_output", "合格数": "qualified_count", "录入人": "recorder",
    }
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                field = field_map.get(str(h), str(h))
                row_dict[field] = row[i]
        if row_dict.get("year"):
            rows_data.append(row_dict)
    wb.close()
    return batch_import_weekly(db, rows_data)


def import_weekly_raw_excel(db: Session, file: UploadFile):
    from openpyxl import load_workbook

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents), read_only=True)
    sheet_names = wb.sheetnames
    ws = wb["Detection_Result"] if "Detection_Result" in sheet_names else wb.active
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    required = ["检测时间", "工站名称", "机型名称", "检测结果"]
    col_idx = {}
    for name in required:
        if name in headers:
            col_idx[name] = headers.index(name)
        else:
            wb.close()
            raise HTTPException(status_code=400, detail=f"缺少必要列: {name}，请确认文件为 DetectionResult 格式")

    grouped = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        det_time = row[col_idx["检测时间"]]
        station = row[col_idx["工站名称"]]
        model = row[col_idx["机型名称"]]
        result = row[col_idx["检测结果"]]
        if not det_time or not station or not model:
            continue
        if isinstance(det_time, datetime):
            dt = det_time
        else:
            time_str = str(det_time).split("~")[0].strip()
            try:
                dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    dt = datetime.strptime(time_str, "%Y/%m/%d %H:%M:%S")
                except ValueError:
                    continue
        iso = dt.isocalendar()
        iso_year, iso_week = iso[0], iso[1]
        line = {"2F3LDK": "3线", "2F4L": "4线", "2F6LDK": "6线", "2F6LCK": "6线"}.get(str(station).strip(), str(station).strip())
        project = {"巴拿马D壳": "巴拿马8139-D壳", "JGLNL336NJ": "机械革命-LNL336", "荣耀D壳-2F": "荣耀D壳-欧亚版", "荣耀C壳_2F": "荣耀C壳-欧亚版"}.get(str(model).strip(), str(model).strip())
        key = (iso_year, iso_week, line, project)
        if key not in grouped:
            grouped[key] = {"total": 0, "ok": 0}
        grouped[key]["total"] += 1
        if str(result).strip() == "OK":
            grouped[key]["ok"] += 1
    wb.close()
    if not grouped:
        raise HTTPException(status_code=400, detail="未解析到有效检测数据")
    rows_data = [{
        "year": year,
        "week_number": week,
        "production_line": line,
        "project": project,
        "total_output": counts["total"],
        "qualified_count": counts["ok"],
        "recorder": "system",
    } for (year, week, line, project), counts in grouped.items()]
    return batch_import_weekly(db, rows_data)


def list_monthly(db: Session, page: int, page_size: int, year: Optional[int] = None, month: Optional[int] = None, project: Optional[str] = None):
    items, total = get_monthly_production_paginated(db, page, page_size, year, month, None, project)
    stats = get_monthly_summary_stats(db, year, month)
    return [monthly_to_dict(item) for item in items], total, stats


def monthly_stats(db: Session, year: Optional[int] = None, month: Optional[int] = None):
    return get_monthly_summary_stats(db, year, month)


def monthly_trend(db: Session, year: Optional[int] = None):
    return get_monthly_trend(db, year)


def create_monthly(db: Session, payload: MonthlyProductionCreate):
    record = create_monthly_production(db, payload.model_dump())
    return monthly_to_dict(record)


def get_monthly_by_id(db: Session, record_id: int):
    from app.core.crud import get_monthly_production_by_id
    record = get_monthly_production_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    return monthly_to_dict(record)


def update_monthly(db: Session, record_id: int, payload: MonthlyProductionUpdate):
    record = update_monthly_production(db, record_id, payload.model_dump(exclude_unset=True))
    if not record:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    return monthly_to_dict(record)


def delete_monthly(db: Session, record_id: int):
    ok = delete_monthly_production(db, record_id)
    if not ok:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    return True


def generate_monthly(db: Session, body: MonthlyGenerateRequest, recorder: str):
    return generate_monthly_from_weekly(db, body.year, body.month, recorder)
