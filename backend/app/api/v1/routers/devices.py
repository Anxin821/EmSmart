from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.auth import get_current_user, require_role
from app.schemas import AoiAiDeviceCreate, AoiAiDeviceUpdate, ApiResponse, PaginatedData
from app.services import devices_service as service

router = APIRouter(prefix="/devices", tags=["AOI&AI设备"])


@router.get("")
def list_devices(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, total = service.list_devices(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status, device_type=device_type)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=items))


@router.post("")
def create_device(
    data: AoiAiDeviceCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    device = service.create_device(db, data.model_dump(), request, current_user["username"])
    return ApiResponse(data=device)


@router.put("/{item_id}")
def edit_device(
    item_id: int,
    data: AoiAiDeviceUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    device = service.edit_device(db, item_id, data.model_dump(exclude_unset=True), request, current_user["username"])
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return ApiResponse(data=device)


@router.delete("/{item_id}")
def remove_device(
    item_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    ok = service.remove_device(db, item_id, request, current_user["username"])
    if not ok:
        raise HTTPException(status_code=404, detail="设备不存在")
    return ApiResponse(message="删除成功")


@router.get("/detail/{item_id}")
def get_device(
    item_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    device = service.get_device(db, item_id)
    if not device:
        raise HTTPException(status_code=404, detail="设备不存在")
    return ApiResponse(data=device)


@router.post("/import")
def import_devices(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    if not file.filename.endswith((".xlsx", ".xls")):
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    from openpyxl import load_workbook
    wb = load_workbook(filename=BytesIO(contents), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                field_map = {
                    "设备ID": "device_id", "名称": "name", "类型": "device_type",
                    "产线": "production_line", "位置": "location", "IP": "ip_address",
                    "状态": "status", "负责人": "responsible_person", "安装日期": "install_date",
                    "备注": "remark",
                }
                field = field_map.get(str(h), str(h))
                row_dict[field] = row[i]
        if row_dict.get("device_id"):
            rows_data.append(row_dict)
    wb.close()
    count = service.import_devices(db, rows_data, request, current_user["username"])
    return ApiResponse(data={"imported": count}, message=f"成功导入 {count} 条记录")


@router.get("/export/excel")
def export_devices(
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    device_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = service.list_devices(db, 1, 10000, keyword, production_line, status, device_type)
    output = service.export_devices_rows(db, items[0] if isinstance(items, tuple) else items)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=aoi_ai_devices.xlsx"})
