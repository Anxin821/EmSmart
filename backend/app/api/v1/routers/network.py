from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.core.auth import get_current_user, require_role
from app.schemas import (
    ServerCreate, ServerUpdate, AgingRackCreate, AgingRackUpdate,
    WifiApCreate, WifiApUpdate, ApiResponse, PaginatedData,
)
from app.services import network_service as service

router = APIRouter(prefix="/network", tags=["车间网络"])


@router.get("/servers")
def list_servers(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None,
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, total = service.list_servers(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=items))


@router.post("/servers")
def add_server(
    data: ServerCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    server = service.add_server(db, data.model_dump(), request, current_user["username"])
    return ApiResponse(data=server)


@router.put("/servers/{server_id}")
def edit_server(
    server_id: str,
    data: ServerUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    server = service.edit_server(db, server_id, data.model_dump(exclude_unset=True), request, current_user["username"])
    if not server:
        raise HTTPException(status_code=404, detail="服务器不存在")
    return ApiResponse(data=server)


@router.delete("/servers/{server_id}")
def remove_server(
    server_id: str,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    ok = service.remove_server(db, server_id, request, current_user["username"])
    if not ok:
        raise HTTPException(status_code=404, detail="服务器不存在")
    return ApiResponse(message="删除成功")


@router.post("/servers/check-all")
def check_all_servers(
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    res = service.check_all_servers(db, request, current_user["username"])
    return ApiResponse(data=res)


@router.post("/servers/import")
def import_servers(file: UploadFile = File(...), request: Request = None,
                   db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin", "engineer"))):
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls")
    from openpyxl import load_workbook
    wb = load_workbook(filename=BytesIO(file.file.read()), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    field_map = {"服务器ID": "server_id", "名称": "name", "产线": "production_line",
                 "机柜位置": "rack_location", "IP": "ip_address", "型号": "model",
                 "OS": "os", "状态": "status", "CPU使用率": "cpu_usage",
                 "内存使用率": "memory_usage", "硬盘使用率": "disk_usage", "负责人": "responsible_person"}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        r = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                f = field_map.get(str(h), str(h))
                r[f] = row[i]
        if r.get("server_id"):
            rows.append(r)
    wb.close()
    count = service.import_servers(db, rows, request, current_user["username"])
    return ApiResponse(data={"imported": count})


@router.get("/servers/export")
def export_servers(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = service.list_servers(db, 1, 10000, keyword, production_line, status)
    output = service.export_servers_rows(db, items)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=servers.xlsx"})


# aging racks
@router.get("/aging-racks")
def list_aging_racks(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = service.list_aging_racks(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=items))


@router.post("/aging-racks")
def add_aging_rack(data: AgingRackCreate, request: Request, db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin", "engineer"))):
    rack = service.add_aging_rack(db, data.model_dump(), request, current_user["username"])
    return ApiResponse(data=rack)


@router.put("/aging-racks/{rack_id}")
def edit_aging_rack(rack_id: str, data: AgingRackUpdate, request: Request,
                    db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    rack = service.edit_aging_rack(db, rack_id, data.model_dump(exclude_unset=True), request, current_user["username"])
    if not rack: raise HTTPException(404, "老化架不存在")
    return ApiResponse(data=rack)


@router.delete("/aging-racks/{rack_id}")
def remove_aging_rack(rack_id: str, request: Request, db: Session = Depends(get_db),
                      current_user: dict = Depends(require_role("admin"))):
    ok = service.remove_aging_rack(db, rack_id, request, current_user["username"])
    if not ok: raise HTTPException(404, "老化架不存在")
    return ApiResponse(message="删除成功")


@router.get("/aging-racks/export")
def export_aging_racks(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = service.list_aging_racks(db, 1, 10000, keyword, production_line, status)
    output = service.export_aging_racks_rows(db, items)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=aging_racks.xlsx"})


# wifi aps
@router.get("/wifi-aps")
def list_wifi_aps(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, production_line: Optional[str] = None,
    status: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = service.list_wifi_aps(db, page=page, page_size=page_size, keyword=keyword, production_line=production_line, status=status)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=items))


@router.post("/wifi-aps")
def add_wifi_ap(data: WifiApCreate, request: Request, db: Session = Depends(get_db),
                current_user: dict = Depends(require_role("admin", "engineer"))):
    ap = service.add_wifi_ap(db, data.model_dump(), request, current_user["username"])
    return ApiResponse(data=ap)


@router.put("/wifi-aps/{ap_id}")
def edit_wifi_ap(ap_id: str, data: WifiApUpdate, request: Request,
                 db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    ap = service.edit_wifi_ap(db, ap_id, data.model_dump(exclude_unset=True), request, current_user["username"])
    if not ap: raise HTTPException(404, "AP不存在")
    return ApiResponse(data=ap)


@router.delete("/wifi-aps/{ap_id}")
def remove_wifi_ap(ap_id: str, request: Request, db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin"))):
    ok = service.remove_wifi_ap(db, ap_id, request, current_user["username"])
    if not ok: raise HTTPException(404, "AP不存在")
    return ApiResponse(message="删除成功")


@router.get("/wifi-aps/export")
def export_wifi_aps(
    production_line: Optional[str] = None,
    status: Optional[str] = None,
    keyword: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = service.list_wifi_aps(db, 1, 10000, keyword, production_line, status)
    output = service.export_wifi_aps_rows(db, items)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=wifi_aps.xlsx"})