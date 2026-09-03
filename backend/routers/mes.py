"""
智能工厂工作任务管理平台 - MES 系统路由
提供工单/BUG/需求 CRUD 及状态流转
"""
from io import BytesIO
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from core.database import get_db
from core.auth import get_current_user, require_role, write_operation_log
from schemas import (
    WorkOrderCreate, WorkOrderUpdate,
    BugCreate, BugUpdate,
    DevRequestCreate, DevRequestUpdate,
    StatusUpdateRequest,  # 原文件内联 Schema，已迁入 schemas/mes.py
    ApiResponse, PaginatedData,
)
from core.crud import (
    get_work_orders_paginated, create_work_order, update_work_order, delete_work_order,
    get_bugs_paginated, create_bug, update_bug, delete_bug,
    get_dev_requests_paginated, create_dev_request, update_dev_request, delete_dev_request,
    get_mes_dashboard,
)

router = APIRouter(prefix="/mes", tags=["MES系统"])


# ============================================================
# 看板聚合
# ============================================================
@router.get("/dashboard")
def mes_dashboard(
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """MES 管理看板聚合数据：BUG 修复率、需求完成率、月度状态堆叠、风险 TOP、里程碑"""
    return ApiResponse(data=get_mes_dashboard(db))


# ============================================================
# 工单
# ============================================================
@router.get("/work-orders")
def list_work_orders(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, status: Optional[str] = None,
    priority: Optional[str] = None, order_type: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = get_work_orders_paginated(db, page, page_size, keyword, status, priority, order_type)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=[_order_to_dict(d) for d in items]))


@router.get("/work-orders/{order_number}")
def get_work_order(order_number: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import WorkOrder as WO
    item = db.query(WO).filter(WO.order_number == order_number).first()
    if not item: raise HTTPException(404, "工单不存在")
    return ApiResponse(data=_order_to_dict(item))


@router.post("/work-orders")
def add_work_order(data: WorkOrderCreate, request: Request, db: Session = Depends(get_db),
                   current_user: dict = Depends(require_role("admin", "engineer"))):
    order = create_work_order(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "work_order", data.order_number, f"新增工单: {data.product_name}", request)
    return ApiResponse(data=_order_to_dict(order))


@router.put("/work-orders/{order_number}")
def edit_work_order(order_number: str, data: WorkOrderUpdate, request: Request,
                    db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    order = update_work_order(db, order_number, data.model_dump(exclude_unset=True))
    if not order:
        raise HTTPException(404, "工单不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "work_order", order_number, "更新工单", request)
    return ApiResponse(data=_order_to_dict(order))


@router.delete("/work-orders/{order_number}")
def remove_work_order(order_number: str, request: Request, db: Session = Depends(get_db),
                      current_user: dict = Depends(require_role("admin"))):
    if not delete_work_order(db, order_number):
        raise HTTPException(404, "工单不存在")
    write_operation_log(db, current_user["username"], "DELETE", "work_order", order_number, "删除工单", request)
    return ApiResponse(message="删除成功")


@router.put("/work-orders/{order_number}/status")
def update_order_status(
    order_number: str,
    body: StatusUpdateRequest,
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """工单状态流转"""
    new_status = body.status
    from datetime import datetime
    update_data = {"status": new_status}
    if new_status == "进行中":
        update_data["actual_start"] = datetime.utcnow()
    elif new_status == "已完成":
        update_data["actual_end"] = datetime.utcnow()
    order = update_work_order(db, order_number, update_data)
    if not order:
        raise HTTPException(404, "工单不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "work_order", order_number, f"状态流转 -> {new_status}", request)
    return ApiResponse(data=_order_to_dict(order))


@router.get("/work-orders/export")
def export_work_orders(
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    order_type: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_work_orders_paginated(db, 1, 10000, keyword, status, priority, order_type)
    wb = Workbook(); ws = wb.active; ws.title = "工单"
    ws.append(["工单号","类型","产品名","优先级","计划开始","计划结束","实际开始","实际结束","状态","负责人","描述"])
    for d in items:
        ws.append([d.order_number, d.order_type, d.product_name, d.priority,
                    str(d.planned_start or ""), str(d.planned_end or ""),
                    str(d.actual_start or ""), str(d.actual_end or ""), d.status,
                    d.responsible_person or "", d.description or ""])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=work_orders.xlsx"})


# ============================================================
# BUG
# ============================================================
@router.get("/bugs")
def list_bugs(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, status: Optional[str] = None,
    severity: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = get_bugs_paginated(db, page, page_size, keyword, status, severity)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=[_bug_to_dict(d) for d in items]))


@router.get("/bugs/{bug_id}")
def get_bug(bug_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import Bug as B
    item = db.query(B).filter(B.bug_id == bug_id).first()
    if not item: raise HTTPException(404, "BUG不存在")
    return ApiResponse(data=_bug_to_dict(item))


@router.post("/bugs")
def add_bug(data: BugCreate, request: Request, db: Session = Depends(get_db),
            current_user: dict = Depends(require_role("admin", "engineer"))):
    bug = create_bug(db, data.model_dump(exclude_unset=False))
    write_operation_log(db, current_user["username"], "CREATE", "bug", bug.bug_id, f"新增BUG: {data.title}", request)
    return ApiResponse(data=_bug_to_dict(bug))


@router.put("/bugs/{bug_id}")
def edit_bug(bug_id: str, data: BugUpdate, request: Request, db: Session = Depends(get_db),
             current_user: dict = Depends(require_role("admin", "engineer"))):
    bug = update_bug(db, bug_id, data.model_dump(exclude_unset=True))
    if not bug: raise HTTPException(404, "BUG不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "bug", bug_id, "更新BUG", request)
    return ApiResponse(data=_bug_to_dict(bug))


@router.delete("/bugs/{bug_id}")
def remove_bug(bug_id: str, request: Request, db: Session = Depends(get_db),
               current_user: dict = Depends(require_role("admin"))):
    if not delete_bug(db, bug_id): raise HTTPException(404, "BUG不存在")
    write_operation_log(db, current_user["username"], "DELETE", "bug", bug_id, "删除BUG", request)
    return ApiResponse(message="删除成功")


@router.put("/bugs/{bug_id}/status")
def update_bug_status(
    bug_id: str,
    body: StatusUpdateRequest,
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    valid = ["确认新增", "修复中", "解决关闭"]
    new_status = body.status
    if new_status not in valid:
        raise HTTPException(400, f"无效状态，可选值: {', '.join(valid)}")
    bug = update_bug(db, bug_id, {"status": new_status})
    if not bug: raise HTTPException(404, "BUG不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "bug", bug_id, f"状态流转 -> {new_status}", request)
    return ApiResponse(data=_bug_to_dict(bug))


@router.get("/bugs/export")
def export_bugs(
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    severity: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_bugs_paginated(db, 1, 10000, keyword, status, severity)
    wb = Workbook(); ws = wb.active; ws.title = "BUG"
    ws.append(["BUG_ID","标题","严重等级","模块","状态","发现人","责任人","创建日期","期限","解决方案"])
    for d in items:
        ws.append([d.bug_id, d.title, d.severity, d.module or "", d.status,
                    d.discoverer or "", d.assignee or "", str(d.created_date or ""), str(d.deadline or ""), d.solution or ""])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=bugs.xlsx"})


# ============================================================
# 二次开发需求
# ============================================================
@router.get("/dev-requests")
def list_dev_requests(
    page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100),
    keyword: Optional[str] = None, status: Optional[str] = None,
    priority: Optional[str] = None,
    db: Session = Depends(get_db), current_user: dict = Depends(get_current_user),
):
    items, total = get_dev_requests_paginated(db, page, page_size, keyword, status, priority)
    return ApiResponse(data=PaginatedData(total=total, page=page, page_size=page_size, items=[_req_to_dict(d) for d in items]))


@router.get("/dev-requests/{request_id}")
def get_dev_request(request_id: str, db: Session = Depends(get_db), current_user: dict = Depends(get_current_user)):
    from models import DevRequest as DR
    item = db.query(DR).filter(DR.request_id == request_id).first()
    if not item: raise HTTPException(404, "需求不存在")
    return ApiResponse(data=_req_to_dict(item))


@router.post("/dev-requests")
def add_dev_request(data: DevRequestCreate, request: Request, db: Session = Depends(get_db),
                    current_user: dict = Depends(require_role("admin", "engineer"))):
    req = create_dev_request(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "dev_request", data.request_id, f"新增需求: {data.title}", request)
    return ApiResponse(data=_req_to_dict(req))


@router.put("/dev-requests/{request_id}")
def edit_dev_request(request_id: str, data: DevRequestUpdate, req_ctx: Request,
                     db: Session = Depends(get_db), current_user: dict = Depends(require_role("admin", "engineer"))):
    req = update_dev_request(db, request_id, data.model_dump(exclude_unset=True))
    if not req: raise HTTPException(404, "需求不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "dev_request", request_id, "更新需求", req_ctx)
    return ApiResponse(data=_req_to_dict(req))


@router.delete("/dev-requests/{request_id}")
def remove_dev_request(request_id: str, req_ctx: Request, db: Session = Depends(get_db),
                       current_user: dict = Depends(require_role("admin"))):
    if not delete_dev_request(db, request_id): raise HTTPException(404, "需求不存在")
    write_operation_log(db, current_user["username"], "DELETE", "dev_request", request_id, "删除需求", req_ctx)
    return ApiResponse(message="删除成功")


@router.get("/dev-requests/export")
def export_dev_requests(
    keyword: Optional[str] = None,
    status: Optional[str] = None,
    priority: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_dev_requests_paginated(db, 1, 10000, keyword, status, priority)
    wb = Workbook(); ws = wb.active; ws.title = "需求"
    ws.append(["需求ID","标题","来源","优先级","状态","期望日期","负责人","进度%","描述"])
    for d in items:
        ws.append([d.request_id, d.title, d.source or "", d.priority, d.status,
                    str(d.expected_date or ""), d.responsible_person or "", d.progress, d.description or ""])
    output = BytesIO(); wb.save(output); output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=dev_requests.xlsx"})


# ============================================================
# 辅助函数
# ============================================================
def _order_to_dict(d) -> dict:
    return {
        "id": d.id, "order_number": d.order_number, "order_type": d.order_type,
        "product_name": d.product_name, "priority": d.priority,
        "planned_start": d.planned_start.isoformat() if d.planned_start else None,
        "planned_end": d.planned_end.isoformat() if d.planned_end else None,
        "actual_start": d.actual_start.isoformat() if d.actual_start else None,
        "actual_end": d.actual_end.isoformat() if d.actual_end else None,
        "status": d.status, "responsible_person": d.responsible_person,
        "description": d.description,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _bug_to_dict(d) -> dict:
    return {
        "id": d.id, "bug_id": d.bug_id, "title": d.title,
        "severity": (d.severity or "").strip().lstrip("|").rstrip("|").strip(),
        "module": d.module,
        "status": (d.status or "").strip().lstrip("|").rstrip("|").strip(),
        "discoverer": d.discoverer, "assignee": d.assignee,
        "created_date": d.created_date.isoformat() if d.created_date else None,
        "deadline": d.deadline.isoformat() if d.deadline else None,
        "solution": d.solution,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _req_to_dict(d) -> dict:
    return {
        "id": d.id, "request_id": d.request_id, "title": d.title, "source": d.source,
        "priority": (d.priority or "").strip().lstrip("|").rstrip("|").strip(),
        "status": (d.status or "").strip().lstrip("|").rstrip("|").strip(),
        "submitter": d.submitter, "assignee": d.assignee,
        "expected_date": d.expected_date.isoformat() if d.expected_date else None,
        "responsible_person": d.responsible_person, "progress": d.progress,
        "description": d.description,
        "created_at": d.created_at.isoformat() if d.created_at else None,
        "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }
