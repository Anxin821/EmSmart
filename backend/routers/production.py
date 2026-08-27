"""
智能工厂工作任务管理平台 - 生产数据路由
提供周报/月报 CRUD、批量导入导出、月报汇总生成
"""
from io import BytesIO
from datetime import datetime
from typing import Optional
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Request
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook, load_workbook

from core.database import get_db
from core.auth import get_current_user, require_role, write_operation_log
from schemas import (
    WeeklyProductionCreate, WeeklyProductionUpdate,
    MonthlyProductionCreate, MonthlyProductionUpdate,
    MonthlyGenerateRequest,  # 原文件内联 Schema，已迁入 schemas/production.py
    ApiResponse, PaginatedData,
)
from core.crud import (
    get_weekly_production_paginated, get_weekly_production_by_id,
    create_weekly_production, update_weekly_production,
    delete_weekly_production,
    get_monthly_production_paginated, get_monthly_summary_stats, get_monthly_trend,
    create_monthly_production, update_monthly_production,
    delete_monthly_production, generate_monthly_from_weekly, batch_import_weekly,
)

router = APIRouter(prefix="/production", tags=["生产数据"])


# ============================================================
# 周报
# ============================================================
@router.get("/weekly")
def list_weekly(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    year: Optional[int] = None,
    week: Optional[int] = None,
    production_line: Optional[str] = None,
    project: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取周报分页列表"""
    items, total = get_weekly_production_paginated(db, page, page_size, year, week, production_line, project)
    return ApiResponse(data=PaginatedData(
        total=total, page=page, page_size=page_size,
        items=[_weekly_to_dict(d) for d in items],
    ))


@router.post("/weekly")
def create_weekly(
    data: WeeklyProductionCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """录入周报数据"""
    record = create_weekly_production(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "weekly_production", str(record.id), f"录入周报: {data.year}W{data.week_number}", request)
    return ApiResponse(data=_weekly_to_dict(record))


@router.get("/weekly/{record_id}")
def get_weekly(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取单条周报（编辑时回填用）"""
    from core.crud import get_weekly_production_by_id
    record = get_weekly_production_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    return ApiResponse(data=_weekly_to_dict(record))


@router.put("/weekly/{record_id}")
def edit_weekly(
    record_id: int,
    data: WeeklyProductionUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    record = update_weekly_production(db, record_id, data.model_dump(exclude_unset=True))
    if not record:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "weekly_production", str(record_id), "更新周报", request)
    return ApiResponse(data=_weekly_to_dict(record))


@router.delete("/weekly/{record_id}")
def remove_weekly(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    ok = delete_weekly_production(db, record_id)
    if not ok:
        raise HTTPException(status_code=404, detail="周报记录不存在")
    write_operation_log(db, current_user["username"], "DELETE", "weekly_production", str(record_id), "删除周报", request)
    return ApiResponse(message="删除成功")


@router.post("/weekly/import")
def import_weekly(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """批量导入周报（Excel）"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents), read_only=True)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    field_map = {
        "年": "year", "周数": "week_number", "产线": "production_line", "项目": "project",
        "总产量": "total_output", "合格数": "qualified_count", "录入人": "recorder",
    }

    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                field = field_map.get(str(h), str(h))
                row_dict[field] = row[i]
        if row_dict.get("year"):
            rows_data.append(row_dict)
    wb.close()

    count = batch_import_weekly(db, rows_data)
    write_operation_log(db, current_user["username"], "CREATE", "weekly_production", None, f"批量导入 {count} 条周报", request)
    return ApiResponse(data={"imported": count}, message=f"成功导入 {count} 条记录")


# ---------- 原始检测数据批量导入 ----------
# 工站名称 → 产线 映射
STATION_LINE_MAP = {
    "2F3LDK": "3线",
    "2F4L": "4线",
    "2F6LDK": "6线",
    "2F6LCK": "6线",
}

# 机型名称 → 项目编码 映射
MODEL_PROJECT_MAP = {
    "巴拿马D壳": "巴拿马8139-D壳",
    "JGLNL336NJ": "机械革命-LNL336",
    "荣耀D壳-2F": "荣耀D壳-欧亚版",
    "荣耀C壳_2F": "荣耀C壳-欧亚版",
}


@router.post("/weekly/import-raw")
def import_weekly_raw(
    file: UploadFile = File(...),
    request: Request = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """批量导入原始检测数据（DetectionResult Excel），自动按周汇总后写入周报"""
    if not file.filename.endswith(('.xlsx', '.xls')):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xls 文件")

    contents = file.file.read()
    wb = load_workbook(filename=BytesIO(contents), read_only=True)

    # 尝试读取 Detection_Result sheet，不存在则用 active sheet
    sheet_names = wb.sheetnames
    ws = wb["Detection_Result"] if "Detection_Result" in sheet_names else wb.active

    # 读取表头，定位所需列
    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
    required = ["检测时间", "工站名称", "机型名称", "检测结果"]
    col_idx = {}
    for col_name in required:
        if col_name in headers:
            col_idx[col_name] = headers.index(col_name)
        else:
            wb.close()
            raise HTTPException(status_code=400, detail=f"缺少必要列: {col_name}，请确认文件为 DetectionResult 格式")

    # 逐行读取并分组聚合
    grouped = {}  # key: (iso_year, iso_week, line, project) → {total, ok}
    for row in ws.iter_rows(min_row=2, values_only=True):
        det_time = row[col_idx["检测时间"]]
        station  = row[col_idx["工站名称"]]
        model    = row[col_idx["机型名称"]]
        result   = row[col_idx["检测结果"]]

        if not det_time or not station or not model:
            continue

        # 解析检测时间（取开始时间，格式 "2025-08-20 08:00:00~..." 或直接 datetime）
        if isinstance(det_time, datetime):
            dt = det_time
        else:
            time_str = str(det_time).split("~")[0].strip()
            try:
                dt = datetime.strptime(time_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                try:
                    dt = datetime.strptime(time_str, "%Y/%m/%d %H:%M:%S")
                except ValueError:
                    continue

        iso = dt.isocalendar()
        iso_year, iso_week = iso[0], iso[1]

        line    = STATION_LINE_MAP.get(str(station).strip(), str(station).strip())
        project = MODEL_PROJECT_MAP.get(str(model).strip(), str(model).strip())

        key = (iso_year, iso_week, line, project)
        if key not in grouped:
            grouped[key] = {"total": 0, "ok": 0}
        grouped[key]["total"] += 1
        if str(result).strip() == "OK":
            grouped[key]["ok"] += 1

    wb.close()

    if not grouped:
        raise HTTPException(status_code=400, detail="未解析到有效检测数据")

    # 构建周报记录，录入人为当前登录用户
    recorder = current_user["username"]
    rows_data = []
    for (year, week, line, project), counts in grouped.items():
        rows_data.append({
            "year": year,
            "week_number": week,
            "production_line": line,
            "project": project,
            "total_output": counts["total"],
            "qualified_count": counts["ok"],
            "recorder": recorder,
        })

    count = batch_import_weekly(db, rows_data)
    write_operation_log(
        db, current_user["username"], "CREATE", "weekly_production",
        None, f"原始检测数据批量导入 {count} 条周报", request
    )
    return ApiResponse(data={"imported": count}, message=f"成功导入 {count} 条周报记录")


@router.get("/weekly/export")
def export_weekly(
    year: Optional[int] = None,
    week: Optional[int] = None,
    production_line: Optional[str] = None,
    project: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """导出周报为 Excel"""
    items, _ = get_weekly_production_paginated(db, 1, 10000, year, week, production_line, project)
    wb = Workbook()
    ws = wb.active
    ws.title = "周报"
    headers = ["年", "周数", "产线", "项目", "总产量", "合格数", "不良数", "直通率(%)", "录入人", "更新时间"]
    ws.append(headers)
    for d in items:
        ws.append([d.year, d.week_number, d.production_line, d.project, d.total_output,
                    d.qualified_count, d.defect_count or 0, d.yield_rate or 0,
                    d.recorder or "", str(d.updated_at)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=weekly_production.xlsx"})


# ============================================================
# 月报
# ============================================================
@router.get("/monthly")
def list_monthly(
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    year: Optional[int] = None,
    month: Optional[int] = None,
    project: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, total = get_monthly_production_paginated(db, page, page_size, year, month, None, project)
    stats = get_monthly_summary_stats(db, year, month)
    return ApiResponse(data=PaginatedData(
        total=total, page=page, page_size=page_size,
        items=[_monthly_to_dict(d) for d in items],
        extra=stats,
    ))


@router.get("/monthly/stats")
def monthly_stats(
    year: Optional[int] = None,
    month: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """月报总数统计：总产量/总合格数/总直通率，可按年/月筛选"""
    stats = get_monthly_summary_stats(db, year, month)
    return ApiResponse(data=stats)


@router.get("/monthly/trend")
def monthly_trend(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """按月（年,月）聚合的产量 + 直通率趋势，用于看板折线/柱状图。
    口径与列表页统一：sum(monthly_total_output)、sum(monthly_qualified_count)，直通率由两者计算。"""
    items = get_monthly_trend(db, year)
    return ApiResponse(data={"items": items})


@router.post("/monthly")
def create_monthly(
    data: MonthlyProductionCreate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    record = create_monthly_production(db, data.model_dump())
    write_operation_log(db, current_user["username"], "CREATE", "monthly_production", str(record.id), f"录入月报: {data.year}M{data.month}", request)
    return ApiResponse(data=_monthly_to_dict(record))


@router.get("/monthly/{record_id}")
def get_monthly(
    record_id: int,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    """获取单条月报（编辑时回填用）"""
    from core.crud import get_monthly_production_by_id
    record = get_monthly_production_by_id(db, record_id)
    if not record:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    return ApiResponse(data=_monthly_to_dict(record))


@router.put("/monthly/{record_id}")
def edit_monthly(
    record_id: int,
    data: MonthlyProductionUpdate,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    record = update_monthly_production(db, record_id, data.model_dump(exclude_unset=True))
    if not record:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    write_operation_log(db, current_user["username"], "UPDATE", "monthly_production", str(record_id), "更新月报", request)
    return ApiResponse(data=_monthly_to_dict(record))


@router.delete("/monthly/{record_id}")
def remove_monthly(
    record_id: int,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin")),
):
    ok = delete_monthly_production(db, record_id)
    if not ok:
        raise HTTPException(status_code=404, detail="月报记录不存在")
    write_operation_log(db, current_user["username"], "DELETE", "monthly_production", str(record_id), "删除月报", request)
    return ApiResponse(message="删除成功")


@router.post("/monthly/generate")
def generate_monthly(
    body: MonthlyGenerateRequest,
    request: Request,
    db: Session = Depends(get_db),
    current_user: dict = Depends(require_role("admin", "engineer")),
):
    """从周报汇总生成月报（需权限）"""
    recorder = current_user["username"]
    count = generate_monthly_from_weekly(db, body.year, body.month, recorder)
    write_operation_log(db, current_user["username"], "CREATE", "monthly_production", None, f"汇总生成 {body.year}M{body.month} 月报，共 {count} 条", request)
    return ApiResponse(data={"generated": count}, message=f"成功生成 {count} 条月报记录")


@router.get("/monthly/export")
def export_monthly(
    year: Optional[int] = None,
    month: Optional[int] = None,
    production_line: Optional[str] = None,
    project: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: dict = Depends(get_current_user),
):
    items, _ = get_monthly_production_paginated(db, 1, 10000, year, month, production_line, project)
    wb = Workbook()
    ws = wb.active
    ws.title = "月报"
    headers = ["年", "月", "产线", "项目", "月总产量", "月合格数", "月不良数", "月直通率(%)", "录入人", "更新时间"]
    ws.append(headers)
    for d in items:
        ws.append([d.year, d.month, d.production_line, d.project, d.monthly_total_output,
                    d.monthly_qualified_count, d.monthly_defect_count or 0, d.monthly_yield_rate or 0,
                    d.recorder or "", str(d.updated_at)])
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": "attachment; filename=monthly_production.xlsx"})


# ============================================================
# 辅助函数
# ============================================================
def _weekly_to_dict(d) -> dict:
    return {
        "id": d.id, "year": d.year, "week_number": d.week_number,
        "production_line": d.production_line, "project": d.project,
        "total_output": d.total_output, "qualified_count": d.qualified_count,
        "defect_count": d.defect_count, "yield_rate": d.yield_rate,
        "recorder": d.recorder, "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }


def _monthly_to_dict(d) -> dict:
    return {
        "id": d.id, "year": d.year, "month": d.month,
        "production_line": d.production_line, "project": d.project,
        "monthly_total_output": d.monthly_total_output,
        "monthly_qualified_count": d.monthly_qualified_count,
        "monthly_defect_count": d.monthly_defect_count,
        "monthly_yield_rate": d.monthly_yield_rate,
        "recorder": d.recorder, "updated_at": d.updated_at.isoformat() if d.updated_at else None,
    }
